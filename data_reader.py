import openpyxl
from jinja2 import Template




class FieldMissingError(Exception):
    pass


def get_gen(sheet, cfg_sheet):

    sheet = sheet.values
    headers = next(sheet)

    if headers[-1] != 'Status':
        raise FieldMissingError("Status Field Is Missing In Excel Sheet")

    data = sheet
    cfg_template = Template(cfg_sheet['A1'].value)
    idx = 0
    for row in data:

        var_dict = {k: v for k, v in zip(headers, row)}
        cfg = cfg_template.render(**var_dict)
        conf = [line.strip() for line in cfg.split('\n') if line.strip()]
        task_data = {
            "task_id": idx,
            "ipaddr": row[1],
            "model": row[2],
            "status": row[-1],
            "commands": conf
        }
        yield task_data
        idx += 1


def get_wbook(file_name):

    wb = openpyxl.load_workbook(file_name, data_only=True)

    sheet = wb['cfgdata']
    cfg_sheet = wb['cfgtemplate']

    data_len = sheet.max_row - 1
    st = chr(64 + sheet.max_column)
    return wb, st, data_len, get_gen(sheet, cfg_sheet)




# wb, datalen, gen = get_wbook(FILE_NAME)
#
# print(datalen)
# i=0
# while i<100:
#     print(next(gen))
#     i += 1